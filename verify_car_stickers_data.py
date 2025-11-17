#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
تحقق من بيانات ملصقات السيارات
Verify Car Stickers Data
"""

import openpyxl
import json
from datetime import datetime
from collections import defaultdict, Counter

def analyze_car_stickers(excel_file):
    """تحليل شامل لبيانات ملصقات السيارات"""
    
    wb = openpyxl.load_workbook(excel_file, read_only=True)
    
    results = {
        'تاريخ_التحليل': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'ملف_البيانات': excel_file,
        'الأوراق': {},
        'إحصائيات_عامة': {},
        'تفاصيل_الملصقات': {}
    }
    
    total_active = 0
    total_cancelled = 0
    
    # تحليل كل ورقة
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # الحصول على العناوين
        headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]
        
        # عد الصفوف غير الفارغة
        data_rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if any(cell is not None for cell in row):
                data_rows.append(row)
        
        num_rows = len(data_rows)
        
        # إحصائيات للورقة الحالية
        sheet_stats = {
            'عدد_الأعمدة': len(headers),
            'العناوين': headers,
            'عدد_الصفوف': num_rows,
            'إحصائيات_تفصيلية': {}
        }
        
        # تحليل البيانات
        if sheet_name == 'فعال':
            total_active = num_rows
            sheet_stats['الحالة'] = 'ملصقات فعالة'
            
            # تحليل إضافي للملصقات الفعالة
            buildings = Counter()
            units = Counter()
            vehicle_types = Counter()
            
            for row in data_rows:
                if len(row) >= 10:
                    # المبنى
                    if row[8] is not None:
                        buildings[str(row[8])] += 1
                    # الوحدة
                    if row[7] is not None:
                        units[str(row[7])] += 1
                    # نوع المركبة (محاولة استخراج الماركة)
                    if row[5] is not None:
                        vehicle_info = str(row[5]).split()[0] if row[5] else "غير محدد"
                        vehicle_types[vehicle_info] += 1
            
            sheet_stats['إحصائيات_تفصيلية'] = {
                'عدد_المباني': len(buildings),
                'عدد_الوحدات': len(units),
                'أكثر_5_مباني': dict(buildings.most_common(5)),
                'أكثر_5_وحدات': dict(units.most_common(5)),
                'أنواع_المركبات': dict(vehicle_types.most_common(10))
            }
            
        elif sheet_name == 'ملغي':
            total_cancelled = num_rows
            sheet_stats['الحالة'] = 'ملصقات ملغية'
        
        results['الأوراق'][sheet_name] = sheet_stats
    
    # الإحصائيات العامة
    results['إحصائيات_عامة'] = {
        'إجمالي_الملصقات_الفعالة': total_active,
        'إجمالي_الملصقات_الملغية': total_cancelled,
        'إجمالي_كل_الملصقات': total_active + total_cancelled,
        'نسبة_الفعالة': round((total_active / (total_active + total_cancelled) * 100), 2) if (total_active + total_cancelled) > 0 else 0,
        'نسبة_الملغية': round((total_cancelled / (total_active + total_cancelled) * 100), 2) if (total_active + total_cancelled) > 0 else 0
    }
    
    wb.close()
    
    return results

def print_arabic_report(results):
    """طباعة تقرير بالعربية"""
    print("=" * 100)
    print("تقرير تحليل بيانات ملصقات السيارات")
    print("=" * 100)
    print(f"\nتاريخ التحليل: {results['تاريخ_التحليل']}")
    print(f"ملف البيانات: {results['ملف_البيانات']}")
    
    print("\n" + "=" * 100)
    print("الإحصائيات العامة")
    print("=" * 100)
    stats = results['إحصائيات_عامة']
    print(f"✅ إجمالي الملصقات الفعالة: {stats['إجمالي_الملصقات_الفعالة']:,}")
    print(f"❌ إجمالي الملصقات الملغية: {stats['إجمالي_الملصقات_الملغية']:,}")
    print(f"📊 إجمالي كل الملصقات: {stats['إجمالي_كل_الملصقات']:,}")
    print(f"📈 نسبة الملصقات الفعالة: {stats['نسبة_الفعالة']}%")
    print(f"📉 نسبة الملصقات الملغية: {stats['نسبة_الملغية']}%")
    
    # تفاصيل كل ورقة
    for sheet_name, sheet_data in results['الأوراق'].items():
        print("\n" + "=" * 100)
        print(f"ورقة: {sheet_name}")
        print("=" * 100)
        print(f"الحالة: {sheet_data['الحالة']}")
        print(f"عدد الصفوف: {sheet_data['عدد_الصفوف']:,}")
        print(f"عدد الأعمدة: {sheet_data['عدد_الأعمدة']}")
        print(f"\nالعناوين:")
        for i, header in enumerate(sheet_data['العناوين'], 1):
            print(f"  {i}. {header}")
        
        if 'إحصائيات_تفصيلية' in sheet_data and sheet_data['إحصائيات_تفصيلية']:
            details = sheet_data['إحصائيات_تفصيلية']
            print(f"\nالإحصائيات التفصيلية:")
            print(f"  • عدد المباني المختلفة: {details.get('عدد_المباني', 0)}")
            print(f"  • عدد الوحدات المختلفة: {details.get('عدد_الوحدات', 0)}")
            
            if 'أكثر_5_مباني' in details:
                print(f"\n  أكثر 5 مباني استخداماً:")
                for building, count in details['أكثر_5_مباني'].items():
                    print(f"    - المبنى {building}: {count} ملصق")
            
            if 'أكثر_5_وحدات' in details:
                print(f"\n  أكثر 5 وحدات استخداماً:")
                for unit, count in details['أكثر_5_وحدات'].items():
                    print(f"    - الوحدة {unit}: {count} ملصق")
            
            if 'أنواع_المركبات' in details:
                print(f"\n  أكثر أنواع المركبات شيوعاً:")
                for vehicle, count in list(details['أنواع_المركبات'].items())[:10]:
                    print(f"    - {vehicle}: {count} مركبة")
    
    print("\n" + "=" * 100)
    print("✅ تم التحليل بنجاح")
    print("=" * 100)

def save_json_report(results, output_file):
    """حفظ التقرير بصيغة JSON"""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\n✅ تم حفظ التقرير في: {output_file}")

if __name__ == '__main__':
    excel_file = 'ملصقات السيارات.xlsx'
    
    print("جاري تحليل بيانات ملصقات السيارات...\n")
    
    try:
        results = analyze_car_stickers(excel_file)
        print_arabic_report(results)
        
        # حفظ النتائج
        json_file = 'car_stickers_analysis.json'
        save_json_report(results, json_file)
        
    except Exception as e:
        print(f"❌ خطأ في التحليل: {str(e)}")
        import traceback
        traceback.print_exc()
