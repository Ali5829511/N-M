#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ù‚Ø§Ø¹Ø¯Ø© Ø¨ÙŠØ§Ù†Ø§Øª Ù…Ù„ØµÙ‚Ø§Øª Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª
Check Car Stickers Database

ÙŠÙ‚ÙˆÙ… Ù‡Ø°Ø§ Ø§Ù„Ø³ÙƒØ±ÙŠØ¨Øª Ø¨Ù€:
1. Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† ÙˆØ¬ÙˆØ¯ Ù…Ù„Ù Excel Ù„Ù„Ù…Ù„ØµÙ‚Ø§Øª
2. Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª
3. Ù…Ù‚Ø§Ø±Ù†Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø¨ÙŠÙ† Ø§Ù„Ù…ØµØ¯Ø±ÙŠÙ†
4. ØªÙ‚Ø¯ÙŠÙ… ØªÙ‚Ø±ÙŠØ± Ø´Ø§Ù…Ù„

This script:
1. Checks for the Excel file with stickers data
2. Verifies the stickers table in the database
3. Compares data between sources
4. Provides comprehensive report
"""

import os
import sys
import json
import traceback
from datetime import datetime

# Constants
EXCEL_FILE = 'Ù…Ù„ØµÙ‚Ø§Øª Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª.xlsx'
JSON_REPORT = 'car_stickers_analysis.json'
VERIFICATION_SCRIPT = 'verify_car_stickers_data.py'
SCHEMA_FILE = 'database/schema.sql'
ENV_FILE = '.env'

def check_excel_file():
    """
    Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† ÙˆØ¬ÙˆØ¯ Ù…Ù„Ù Excel
    Check for Excel file existence and analyze its contents
    """
    excel_file = EXCEL_FILE
    
    print("\n" + "="*80)
    print("ğŸ“„ Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ù…Ù„Ù Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø£ØµÙ„ÙŠ / Checking Source Data File")
    print("="*80)
    
    if os.path.exists(excel_file):
        file_size = os.path.getsize(excel_file)
        file_size_kb = file_size / 1024
        print(f"âœ… Ù…Ù„Ù Excel Ù…ÙˆØ¬ÙˆØ¯: {excel_file}")
        print(f"   Ø§Ù„Ø­Ø¬Ù…: {file_size_kb:.2f} KB")
        
        # Try to load and analyze
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            sheets = wb.sheetnames
            print(f"   Ø§Ù„Ø£ÙˆØ±Ø§Ù‚: {', '.join(sheets)}")
            
            # Count rows in each sheet
            total_rows = 0
            for sheet_name in sheets:
                ws = wb[sheet_name]
                row_count = sum(1 for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row))
                total_rows += row_count
                print(f"   â€¢ {sheet_name}: {row_count:,} ØµÙ")
            
            wb.close()
            print(f"   ğŸ“Š Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª: {total_rows:,}")
            return True, total_rows
            
        except ImportError:
            print("   âš ï¸  Ù…ÙƒØªØ¨Ø© openpyxl ØºÙŠØ± Ù…ØªÙˆÙØ±Ø© - ÙŠÙ…ÙƒÙ† ØªØ«Ø¨ÙŠØªÙ‡Ø§ Ø¨Ù€: pip install openpyxl")
            return True, 0
        except Exception as e:
            print(f"   âš ï¸  Ø®Ø·Ø£ ÙÙŠ Ù‚Ø±Ø§Ø¡Ø© Ø§Ù„Ù…Ù„Ù: {str(e)}")
            return True, 0
    else:
        print(f"âŒ Ù…Ù„Ù Excel ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯: {excel_file}")
        return False, 0

def check_json_analysis():
    """
    Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ù…Ù„Ù ØªØ­Ù„ÙŠÙ„ JSON
    Check for JSON analysis report
    """
    json_file = JSON_REPORT
    
    print("\n" + "="*80)
    print("ğŸ“Š Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† ØªÙ‚Ø±ÙŠØ± Ø§Ù„ØªØ­Ù„ÙŠÙ„ / Checking Analysis Report")
    print("="*80)
    
    if os.path.exists(json_file):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            print(f"âœ… ØªÙ‚Ø±ÙŠØ± Ø§Ù„ØªØ­Ù„ÙŠÙ„ Ù…ÙˆØ¬ÙˆØ¯: {json_file}")
            print(f"   ØªØ§Ø±ÙŠØ® Ø§Ù„ØªØ­Ù„ÙŠÙ„: {data.get('ØªØ§Ø±ÙŠØ®_Ø§Ù„ØªØ­Ù„ÙŠÙ„', 'ØºÙŠØ± Ù…Ø­Ø¯Ø¯')}")
            
            stats = data.get('Ø¥Ø­ØµØ§Ø¦ÙŠØ§Øª_Ø¹Ø§Ù…Ø©', {})
            print(f"\n   Ø§Ù„Ø¥Ø­ØµØ§Ø¦ÙŠØ§Øª Ø§Ù„Ø¹Ø§Ù…Ø©:")
            print(f"   â€¢ Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª Ø§Ù„ÙØ¹Ø§Ù„Ø©: {stats.get('Ø¥Ø¬Ù…Ø§Ù„ÙŠ_Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª_Ø§Ù„ÙØ¹Ø§Ù„Ø©', 0):,}")
            print(f"   â€¢ Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª Ø§Ù„Ù…Ù„ØºÙŠØ©: {stats.get('Ø¥Ø¬Ù…Ø§Ù„ÙŠ_Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª_Ø§Ù„Ù…Ù„ØºÙŠØ©', 0):,}")
            print(f"   â€¢ Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª: {stats.get('Ø¥Ø¬Ù…Ø§Ù„ÙŠ_ÙƒÙ„_Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª', 0):,}")
            
            return True, data
        except Exception as e:
            print(f"   âš ï¸  Ø®Ø·Ø£ ÙÙŠ Ù‚Ø±Ø§Ø¡Ø© Ø§Ù„ØªÙ‚Ø±ÙŠØ±: {str(e)}")
            return True, None
    else:
        print(f"âš ï¸  ØªÙ‚Ø±ÙŠØ± Ø§Ù„ØªØ­Ù„ÙŠÙ„ ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯: {json_file}")
        print(f"   ğŸ’¡ ÙŠÙ…ÙƒÙ†Ùƒ Ø¥Ù†Ø´Ø§Ø¡Ù‡ Ø¨ØªØ´ØºÙŠÙ„: python {VERIFICATION_SCRIPT}")
        return False, None

def check_database_schema():
    """
    Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† ÙˆØ¬ÙˆØ¯ schema Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª
    Check database schema file
    """
    schema_file = SCHEMA_FILE
    
    print("\n" + "="*80)
    print("ğŸ—„ï¸  Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ù‡ÙŠÙƒÙ„ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª / Checking Database Schema")
    print("="*80)
    
    if os.path.exists(schema_file):
        print(f"âœ… Ù…Ù„Ù Schema Ù…ÙˆØ¬ÙˆØ¯: {schema_file}")
        
        try:
            with open(schema_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Check for stickers table
            if 'CREATE TABLE' in content and 'stickers' in content.lower():
                print("   âœ… Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª (stickers) Ù…Ø¹Ø±Ù‘Ù ÙÙŠ Schema")
                
                # Extract column info
                if 'sticker_number' in content:
                    print("   â€¢ ÙŠØ­ØªÙˆÙŠ Ø¹Ù„Ù‰: Ø±Ù‚Ù… Ø§Ù„Ù…Ù„ØµÙ‚ (sticker_number)")
                if 'plate_number' in content:
                    print("   â€¢ ÙŠØ­ØªÙˆÙŠ Ø¹Ù„Ù‰: Ø±Ù‚Ù… Ø§Ù„Ù„ÙˆØ­Ø© (plate_number)")
                if 'owner_name' in content:
                    print("   â€¢ ÙŠØ­ØªÙˆÙŠ Ø¹Ù„Ù‰: Ø§Ø³Ù… Ø§Ù„Ù…Ø§Ù„Ùƒ (owner_name)")
                if 'status' in content:
                    print("   â€¢ ÙŠØ­ØªÙˆÙŠ Ø¹Ù„Ù‰: Ø§Ù„Ø­Ø§Ù„Ø© (status)")
                
                return True
            else:
                print("   âŒ Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯ ÙÙŠ Schema")
                return False
                
        except Exception as e:
            print(f"   âš ï¸  Ø®Ø·Ø£ ÙÙŠ Ù‚Ø±Ø§Ø¡Ø© Schema: {str(e)}")
            return False
    else:
        print(f"âŒ Ù…Ù„Ù Schema ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯: {schema_file}")
        return False

def check_database_connection():
    """
    Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ø§Ù„Ø§ØªØµØ§Ù„ Ø¨Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª
    Check database connection configuration
    """
    print("\n" + "="*80)
    print("ğŸ”Œ Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ø§Ù„Ø§ØªØµØ§Ù„ Ø¨Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª / Checking Database Connection")
    print("="*80)
    
    # Check for .env file
    env_file = ENV_FILE
    if os.path.exists(env_file):
        print(f"âœ… Ù…Ù„Ù Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ù…ÙˆØ¬ÙˆØ¯: {env_file}")
        try:
            with open(env_file, 'r') as f:
                content = f.read()
            if 'DATABASE_URL' in content or 'NETLIFY_DATABASE_URL' in content:
                print("   âœ… Ù…ØªØºÙŠØ±Ø§Øª Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ù…ÙˆØ¬ÙˆØ¯Ø©")
                print("   ğŸ’¡ Ù„Ù„Ø§ØªØµØ§Ù„ Ø¨Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§ØªØŒ Ø§Ø³ØªØ®Ø¯Ù…: node.js Ø£Ùˆ python Ù…Ø¹ Ø§Ù„Ù…ÙƒØªØ¨Ø§Øª Ø§Ù„Ù…Ù†Ø§Ø³Ø¨Ø©")
                return True
            else:
                print("   âš ï¸  Ù…ØªØºÙŠØ±Ø§Øª Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©")
                return False
        except Exception as e:
            print(f"   âš ï¸  Ø®Ø·Ø£ ÙÙŠ Ù‚Ø±Ø§Ø¡Ø© Ù…Ù„Ù Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª: {str(e)}")
            return False
    else:
        print(f"âš ï¸  Ù…Ù„Ù Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯: {env_file}")
        print(f"   ğŸ’¡ Ø±Ø§Ø¬Ø¹: .env.example Ù„Ù„Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§Ù„Ù…Ø·Ù„ÙˆØ¨Ø©")
        return False

def generate_report(excel_exists, excel_rows, json_exists, json_data, schema_exists, db_config_exists):
    """
    Ø¥Ù†Ø´Ø§Ø¡ ØªÙ‚Ø±ÙŠØ± Ø´Ø§Ù…Ù„
    Generate comprehensive report based on all checks
    """
    print("\n" + "="*80)
    print("ğŸ“‹ Ø§Ù„ØªÙ‚Ø±ÙŠØ± Ø§Ù„Ø´Ø§Ù…Ù„ / Comprehensive Report")
    print("="*80)
    print(f"\nØªØ§Ø±ÙŠØ® Ø§Ù„ØªÙ‚Ø±ÙŠØ±: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    print("\n1ï¸âƒ£  Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø£ØµÙ„ÙŠØ© (Excel):")
    if excel_exists:
        print(f"   âœ… Ù…ÙˆØ¬ÙˆØ¯ - {excel_rows:,} Ù…Ù„ØµÙ‚")
    else:
        print("   âŒ ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯")
    
    print("\n2ï¸âƒ£  ØªØ­Ù„ÙŠÙ„ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª (JSON):")
    if json_exists and json_data:
        stats = json_data.get('Ø¥Ø­ØµØ§Ø¦ÙŠØ§Øª_Ø¹Ø§Ù…Ø©', {})
        print(f"   âœ… Ù…ÙˆØ¬ÙˆØ¯ - {stats.get('Ø¥Ø¬Ù…Ø§Ù„ÙŠ_ÙƒÙ„_Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª', 0):,} Ù…Ù„ØµÙ‚")
        print(f"   â€¢ ÙØ¹Ø§Ù„: {stats.get('Ø¥Ø¬Ù…Ø§Ù„ÙŠ_Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª_Ø§Ù„ÙØ¹Ø§Ù„Ø©', 0):,}")
        print(f"   â€¢ Ù…Ù„ØºÙŠ: {stats.get('Ø¥Ø¬Ù…Ø§Ù„ÙŠ_Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª_Ø§Ù„Ù…Ù„ØºÙŠØ©', 0):,}")
    else:
        print("   âš ï¸  ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯ Ø£Ùˆ ØºÙŠØ± Ù…Ø­Ø¯Ù‘Ø«")
    
    print("\n3ï¸âƒ£  Ù‡ÙŠÙƒÙ„ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª:")
    if schema_exists:
        print("   âœ… Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª Ù…Ø¹Ø±Ù‘Ù ÙÙŠ Schema")
    else:
        print("   âŒ Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª ØºÙŠØ± Ù…Ø¹Ø±Ù‘Ù")
    
    print("\n4ï¸âƒ£  Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§Ù„Ø§ØªØµØ§Ù„ Ø¨Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª:")
    if db_config_exists:
        print("   âœ… Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§Ù„Ø§ØªØµØ§Ù„ Ù…ÙˆØ¬ÙˆØ¯Ø©")
    else:
        print("   âš ï¸  Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§Ù„Ø§ØªØµØ§Ù„ ØºÙŠØ± Ù…ÙƒØªÙ…Ù„Ø©")
    
    # Overall status
    print("\n" + "="*80)
    print("ğŸ¯ Ø§Ù„Ø­Ø§Ù„Ø© Ø§Ù„Ø¹Ø§Ù…Ø© / Overall Status")
    print("="*80)
    
    if excel_exists and schema_exists:
        print("\nâœ… **Ø¨ÙŠØ§Ù†Ø§Øª Ù…Ù„ØµÙ‚Ø§Øª Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª Ù…ÙˆØ¬ÙˆØ¯Ø© ÙˆÙ…Ù‡ÙŠØ£Ø©**")
        print("\nğŸ“Œ Ø§Ù„Ø®Ø·ÙˆØ§Øª Ø§Ù„ØªØ§Ù„ÙŠØ©:")
        print("   1. Ø§Ù„ØªØ£ÙƒØ¯ Ù…Ù† Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§Ù„Ø§ØªØµØ§Ù„ Ø¨Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª (.env)")
        print("   2. ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ù…Ù† Excel Ø¥Ù„Ù‰ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª")
        print("   3. Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† ØªØ²Ø§Ù…Ù† Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª")
    elif excel_exists:
        print("\nâš ï¸  **Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ù…ÙˆØ¬ÙˆØ¯Ø© Ù„ÙƒÙ† Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª ØªØ­ØªØ§Ø¬ Ø¥Ø¹Ø¯Ø§Ø¯**")
        print("\nğŸ“Œ Ø§Ù„Ø®Ø·ÙˆØ§Øª Ø§Ù„Ù…Ø·Ù„ÙˆØ¨Ø©:")
        print("   1. Ø¥Ù†Ø´Ø§Ø¡ Ø¬Ø¯ÙˆÙ„ Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª")
        print("   2. ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ù…Ù† Excel")
    else:
        print("\nâŒ **Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù…Ù„ØµÙ‚Ø§Øª ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©**")
        print("\nğŸ“Œ Ø§Ù„Ø®Ø·ÙˆØ§Øª Ø§Ù„Ù…Ø·Ù„ÙˆØ¨Ø©:")
        print("   1. Ø§Ù„Ø­ØµÙˆÙ„ Ø¹Ù„Ù‰ Ù…Ù„Ù Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª (Ù…Ù„ØµÙ‚Ø§Øª Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª.xlsx)")
        print("   2. ØªØ­Ù„ÙŠÙ„ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª")
        print("   3. ØªØ­Ù…ÙŠÙ„Ù‡Ø§ Ø¥Ù„Ù‰ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª")
    
    print("\n" + "="*80)

def main():
    """
    Ø§Ù„Ø¯Ø§Ù„Ø© Ø§Ù„Ø±Ø¦ÙŠØ³ÙŠØ©
    Main function - orchestrates all verification checks
    
    Performs comprehensive verification of:
    - Excel source data file
    - JSON analysis report
    - Database schema
    - Database connection configuration
    
    Generates a detailed report of findings and recommendations.
    """
    print("\n" + "="*80)
    print("ğŸš— Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ù‚Ø§Ø¹Ø¯Ø© Ø¨ÙŠØ§Ù†Ø§Øª Ù…Ù„ØµÙ‚Ø§Øª Ø§Ù„Ø³ÙŠØ§Ø±Ø§Øª")
    print("Car Stickers Database Verification")
    print("="*80)
    
    # Check all components
    excel_exists, excel_rows = check_excel_file()
    json_exists, json_data = check_json_analysis()
    schema_exists = check_database_schema()
    db_config_exists = check_database_connection()
    
    # Generate comprehensive report
    generate_report(excel_exists, excel_rows, json_exists, json_data, schema_exists, db_config_exists)
    
    print("\nâœ… Ø§Ù„ØªØ­Ù‚Ù‚ Ù…ÙƒØªÙ…Ù„ / Verification Complete\n")

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nâš ï¸  ØªÙ… Ø¥ÙŠÙ‚Ø§Ù Ø§Ù„Ø³ÙƒØ±ÙŠØ¨Øª Ù…Ù† Ù‚Ø¨Ù„ Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…")
        sys.exit(0)
    except Exception as e:
        print(f"\nâŒ Ø®Ø·Ø£: {str(e)}")
        traceback.print_exc()
        sys.exit(1)
